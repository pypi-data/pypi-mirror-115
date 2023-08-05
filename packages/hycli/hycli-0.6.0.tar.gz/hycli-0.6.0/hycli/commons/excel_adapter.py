from pathlib import Path
from typing import Dict, AnyStr, NoReturn
import os
from openpyxl.utils import get_column_letter
import pandas as pd
from pandas.io.formats.style import Styler
from hycli.commons.model_comparer import ModelComparer


class ExcelAdapter:
    """Object containing several methods to produce comparison in excel based on extractions in excel

    Args:
        file_path_1: path to hycli extraction in Excel
        file_path_2: path to hycli extraction in Excel

    """

    def __init__(self, file_path_1: Path, file_path_2: Path):
        self.model_1 = self._load_file(file_path_1)
        self.model_2 = self._load_file(file_path_2)
        self.total_difference = None

    @staticmethod
    def _load_file(file_path: Path) -> dict:
        """Loads excel file and parses worksheets into dictionary of key value pairs
        of worksheet name and pandas dataframe.
        """

        def _check_extension() -> str:
            file_name = os.path.basename(file_path)
            extension = os.path.splitext(file_name)[1]
            return extension

        if _check_extension()[:3] == ".xl":
            return dict(pd.read_excel(io=file_path, sheet_name=None, engine="openpyxl"))

        else:
            raise TypeError(
                f"{file_path}: Type {_check_extension()} is currently not supported.\nSupported types: xlsx, xls, "
                f"xlsm"
            )

    def _match_worksheets_by_column_names(self) -> list:
        """

        For the given combination of worksheets in two spreadsheets containing different dataframes match closest one
        based on the column names.

        Returns:
            list: a list of tuples representing worksheet matching
        """
        columns_in_workbook_1 = {
            worksheet_name: df.columns.to_list()
            for worksheet_name, df in self.model_1.items()
        }
        columns_in_workbook_2 = {
            worksheet_name: df.columns.to_list()
            for worksheet_name, df in self.model_2.items()
        }

        def _get_best_match(
                columns_in_df_1: list, propositions: dict, threshold: int = 2
        ):
            """Matches the best combination of column names within a second workbook.

            Args:
                columns_in_df_1: represents a list of columns in first workbook
                propositions: represents a key value pairs of worksheet names and columns of a second workbook
                threshold: represents how many records should be included in second worksheet to consider it
                as a match.

            Returns: string or None

            """
            scores = {}
            for ws_name, columns in propositions.items():
                scores[ws_name] = set(columns_in_df_1).intersection(set(columns))

            if len(max(scores.values())) < threshold:
                return None

            best_match = max(scores, key=scores.get)

            return best_match

        pair_of_matched_worksheets = []
        for ws_name, columns_in_workbook_1 in columns_in_workbook_1.items():
            if columns_in_workbook_2:
                closest_match = _get_best_match(
                    columns_in_workbook_1, columns_in_workbook_2
                )

                if closest_match:
                    pair_of_matched_worksheets.append((ws_name, closest_match))
                    columns_in_workbook_2.pop(closest_match)
            else:
                break

        return pair_of_matched_worksheets

    @staticmethod
    def diff_subtotal(df: pd.DataFrame) -> pd.DataFrame:
        """Provide difference summary for each row"""
        arr = []
        for column in df.columns:
            if column[:5] == "diff_":
                arr.append(column)
        return df[arr].sum(axis=1)

    def summary(self, df_1: pd.DataFrame, df_2: pd.DataFrame) -> pd.DataFrame:
        """Provides the summary of differences, statistics and appends it below comparison dataframe

        Args:
            df_1: ground_truth dataframe
            df_2: extraction dataframe

        Returns: pandas dataframe

        """
        number_of_documents = df_1.shape[0]
        item_categories = len(
            [
                column
                for column in df_2.columns
                if column.startswith("diff_") and column != "diff_total"
            ]
        )
        data_points = number_of_documents * item_categories
        percent_diff = round(self.total_difference / data_points, 2)

        data = {
            1: ["Number of documents", None, number_of_documents],
            2: ["Number of compared columns", None, item_categories],
            3: ["Compared data points", None, data_points],
            4: ["Overall difference in percent", None, percent_diff],
            5: ["Overall accuracy in percent", None, 1 - percent_diff],
        }

        summary = pd.DataFrame.from_dict(
            data, orient="index", columns=["file_name", "distance", "diff_total"]
        )

        df_2 = df_2.append(summary)

        return df_2

    @staticmethod
    def export_comparison_to_excel(
            comparisons: Dict[AnyStr, pd.DataFrame], file_path: Path
    ) -> NoReturn:
        """Exports comparisons to the excel file in given directory.

        Args:
            comparisons (dict): dict containing comparisons stored in pandas dataframes
            file_path (str): output file path of comparison in Excel format

        """

        def _format_excel(df: pd.DataFrame) -> Styler:
            """Takes pandas dataframe and formats excel structure

            Saves dataframes into specific worksheets;
            Sets colours indicating differences;
            Sets column widths.

            Args:
                df: pandas dataframe

            Returns:
                styled pandas dataframe

            """

            def _highlight_records(val):
                val = -1 if val == "" else val

                if val > 0:
                    color = "#FFB6C1"
                    return f"background-color: {color}"

            diff_cols = [col for col in df.columns if "diff_" in col]

            df = (
                df.reset_index(drop=True).style.set_properties(**{"background-color": "#FFFFFF"}).applymap(
                    _highlight_records, subset=list(diff_cols))
            )

            return df

        def _set_column_widths(sheet: str) -> NoReturn:
            """Sets custom column widths for each columns in worksheet"""
            writer.sheets[sheet].column_dimensions["A"].width = 5
            writer.sheets[sheet].column_dimensions["B"].width = 25
            for x in range(3, writer.sheets[sheet].max_column + 1):
                writer.sheets[sheet].column_dimensions[get_column_letter(x)].width = 15

        with pd.ExcelWriter(
                Path(file_path).as_posix(), engine="openpyxl", mode="w"
        ) as writer:
            for k, v in comparisons.items():
                v = _format_excel(v)
                v.to_excel(writer, sheet_name=k)
                _set_column_widths(k)
            writer.save()

    def compare(self) -> Dict[AnyStr, pd.DataFrame]:
        """Compares excel files

        Returns: Comparisons represented as a key value pairs of spreadsheet name and comparison in Pandas Dataframe

        """

        comparisons = {}
        matched_worksheets = self._match_worksheets_by_column_names()

        if matched_worksheets:
            for pair_of_worksheets in matched_worksheets:

                mc = ModelComparer(self.model_1[pair_of_worksheets[0]], self.model_2[pair_of_worksheets[1]])

                def index_duplicated(
                        df: pd.DataFrame, index: str = "file_name"
                ) -> bool:
                    """Checks if index contains duplicated names

                    Args:
                        df: pandas dataframe
                        index: name of matching index which should be matched

                    Returns: True or False

                    """
                    if any(df[index].duplicated().to_list()):
                        return True
                    else:
                        return False

                if any(
                        [
                            index_duplicated(self.model_1[pair_of_worksheets[0]]),
                            index_duplicated(self.model_2[pair_of_worksheets[1]]),
                        ]
                ):
                    comparison_type = "multi"
                else:
                    comparison_type = "single"

                if comparison_type == "single":
                    comparison_results = mc.compare("single")

                else:
                    comparison_results = mc.compare("multi")

                # Add distance columns
                comparison_results = mc.compute_distances(comparison_results)

                # Sort columns
                comparison_results = pd.DataFrame(
                    comparison_results, columns=mc.sort_columns(comparison_results)
                )

                # Prepare total for each row
                comparison_results["diff_total"] = self.diff_subtotal(
                    comparison_results
                )

                # Calculate total difference in percent
                number_of_records = self.model_1[pair_of_worksheets[0]].shape[0]

                # Calculate total difference
                diff_cols = [
                    col for col in comparison_results.columns if ("diff_" in col)
                ]
                sum_of_differences = round(
                    comparison_results[diff_cols].sum(numeric_only=True, axis=0)
                )
                percent_differences = round(sum_of_differences / number_of_records, 2)

                # Place total difference
                comparison_results.loc["diff_total"] = round(sum_of_differences, 2)

                # Place below percent difference
                comparison_results.loc["diff_percent"] = percent_differences

                comparison_results.at["diff_total", "file_name"] = "Total difference"
                comparison_results.at[
                    "diff_percent", "file_name"
                ] = "Total difference in percent"

                self.total_difference = comparison_results.loc[
                    "diff_total", "diff_total"
                ]
                comparison_name = pair_of_worksheets[0]
                comparison_results = self.summary(
                    self.model_1[comparison_name], comparison_results
                )

                if comparison_type == "single":
                    comparison_results.drop(["distance"], axis=1, inplace=True)

                comparisons[comparison_name] = comparison_results

        else:
            raise Exception(
                "The algorithm has no possibility to match any worksheet from second workbook. "
                "Please check the columns names in both files and try again."
            )

        return comparisons
